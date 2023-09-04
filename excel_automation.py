import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle, PatternFill

class ExcelAutomation:
    def __init__(self, task):
        self.task = task

    def start_process(self):
        pass


def read_excel_through_pandas(file_name):
    df = pd.read_excel(file_name)
    print(df)
    return df


def read_excel_using_openpyxl(file_name):
    data = []
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active  # Default First sheet is taken Otherwise we can use sheet = wb["sheet1"]
    # Our sheet has first row as header row and all others as data rows.
    # So while parsing we should start from 1 to max_row ( workbooks indexing starts from 1 instead of 0)
    # this is same as parsing a matrix as Excel sheet is essentially a  2-D matrix.
    for row in range(1, sheet.max_row + 1):
        temp_dict = {}  # This is the dictionary for the row details
        for column in range(1, sheet.max_column + 1):
            temp_dict[sheet.cell(row=1, column=column).value] = sheet.cell(row=row, column=column).value
        data.append(temp_dict)
    print(data)


def write_excel_using_pandas(df):
    df.to_excel("output_pandas.xlsx")


def write_excel_using_openpyxl(data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    if data:
        keys = list(data[0].keys())
        sheet.append(keys)
        for row in data:
            sheet.append(row.values())
    wb.save("openpyxl_simple.xlsx")


def write_excel_with_styling_openpyxl(main_header_style, category_header_style, data):
    pass


# Styles are actually recognized by names for example: we have main_header and category_header two styles.
# These style should be instantiated with same name only once. Will give errors otherwise.
# We are instantiating NamedStyles here.
def main_header_style():
    main_header = NamedStyle(name="main_header")
    main_header.font = Font(bold=True)
    main_header.fill = PatternFill(fgColor="ffcc66", fill_type="solid")
    main_header.border = Border(bottom=Side(border_style="thick"))
    main_header.alignment = Alignment(horizontal="center")
    return main_header


def category_header_style():
    category_header = NamedStyle(name="category_header")
    category_header.fill = PatternFill(fgColor="99ccff", fill_type="solid")
    category_header.font = Font(bold=True)
    category_header.alignment = Alignment(horizontal="center")
    return category_header

if __name__ == "__main__":
    print("This is running as a standalone")
